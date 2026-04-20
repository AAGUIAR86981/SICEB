from datetime import datetime
import io
import csv
import smtplib
import ssl
import os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask import Response, url_for, request
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from config.database import get_db_connection
import json

# Caja de Herramientas: Funciones de ayuda para fechas, correos, archivos y registros del sistema

def dateformat(value, format="%d/%m/%Y"):
    """Convierte una fecha técnica en una fecha fácil de leer para nosotros (ej: 31/12/2024)"""
    if value is None:
        return ""

    if isinstance(value, str):
        # Intentamos adivinar en qué formato viene la fecha para poder transformarla
        try:
            value = datetime.strptime(value, "%Y-%m-%d")
        except ValueError:
            try:
                value = datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return value

    if isinstance(value, datetime):
        return value.strftime(format)

    return value

def from_json(value):
    """Transforma una cadena de texto JSON en una lista o diccionario que el programa pueda entender"""
    if not value:
        return {}
    try:
        return json.loads(value)
    except (ValueError, TypeError):
        return {}

def log_user_activity(user_id, username, activity_type, activity_details):
    """Guarda una nota en la bitácora sobre lo que hizo un usuario (quién, qué y desde dónde)"""
    ip = get_client_ip()
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Si no nos dieron el número de ID directamente, lo buscamos con el nombre de usuario
        if user_id is None and username:
            cursor.execute('SELECT id FROM users WHERE username = %s', (username,))
            user_data = cursor.fetchone()
            if user_data:
                user_id = user_data[0]

        # Escribimos el registro en la tabla de actividades para auditoría
        cursor.execute(
            'INSERT INTO user_activities (user_id, username, activity_type, activity_details, ip_address) VALUES (%s, %s, %s, %s, %s)',
            (user_id, username, activity_type, activity_details, ip)
        )
        conn.commit()
    except Exception as e:
        print(f"Error al intentar guardar la actividad del usuario: {e}")
        if conn: conn.rollback()
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

def exportar_csv(datos_filas, headers, filename, titulo):
    """Crea un archivo CSV sencillo (tipo texto) con los datos que le pasemos"""
    output = io.StringIO()
    writer = csv.writer(output)

    # Agregamos el título y la fecha en que se generó para que el archivo sea profesional
    writer.writerow([titulo])
    writer.writerow([f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    writer.writerow([])
    writer.writerow(headers)

    # Escribimos toda la información fila por fila
    for fila in datos_filas:
        writer.writerow(fila)

    # Preparamos la respuesta para que el navegador del usuario lo descargue automáticamente
    response = Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename={filename}.csv"}
    )
    return response

def exportar_excel_generic(datos_filas, headers, filename, titulo):
    """Crea un archivo de Excel (XLSX) con un diseño bonito (colores, negritas y columnas ajustadas)"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # Colocamos el título principal en la parte superior y lo centramos
        ws.merge_cells(f'A1:{chr(64+len(headers))}1')
        ws['A1'] = titulo
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Ponemos los encabezados de las columnas con un fondo azul claro para que resalten
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        # Empezamos a escribir los datos de los trabajadores o productos a partir de la cuarta fila
        for r_idx, row in enumerate(datos_filas, 4):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Ajustamos el ancho de cada columna automáticamente para que todo el texto sea visible
        from openpyxl.utils import get_column_letter
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column_cells:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length: max_length = length
                except: pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Guardamos el archivo en memoria y lo enviamos al navegador para su descarga
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return Response(
            buffer.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename={filename}.xlsx"}
        )
    except Exception as e:
        print(f"Problema al generar el archivo Excel: {e}")
        return f"Error al generar Excel: {str(e)}", 500

def get_client_ip():
    """Buscamos la dirección IP de la computadora de la persona que está usando el sistema"""
    from flask import request
    if request.headers.get('X-Forwarded-For'):
        # Si estamos detrás de un servidor de red, tomamos la primera dirección real
        return request.headers.get('X-Forwarded-For').split(',')[0]
    return request.remote_addr

def send_reset_email(to_email, token):
    """Envía un correo electrónico especial para que el usuario pueda recuperar su clave olvidada"""
    smtp_server = os.getenv("MAIL_SERVER")
    smtp_port = int(os.getenv("MAIL_PORT", 587))
    smtp_user = os.getenv("MAIL_USERNAME")
    smtp_password = os.getenv("MAIL_PASSWORD")
    sender_email = os.getenv("MAIL_DEFAULT_SENDER", smtp_user)

    # Si no hay un servidor de correos configurado, lo imprimimos por consola (modo desarrollo)
    if not all([smtp_server, smtp_user, smtp_password]):
        print(f"\n[MODO DESARROLLO] LLAVE DE RECUPERACIÓN PARA {to_email}: {token}")
        return True

    # Construimos la dirección web única que el usuario debe clickear
    reset_url = f"{request.host_url.rstrip('/')}{url_for('auth.reset_password_token', token=token)}"

    # Armamos el contenido del correo (en texto plano y en diseño HTML con botón azul)
    message = MIMEMultipart("alternative")
    message["Subject"] = "Recuperación de Contraseña - SICEB"
    message["From"] = sender_email
    message["To"] = to_email

    text = f"Hola,\n\nHemos recibido una solicitud para cambiar tu clave. Haz clic aquí:\n\n{reset_url}\n\nTienes 1 hora para usar este enlace."
    html = f"""
    <html><body>
        <h3>Hola, ¿Olvidaste tu contraseña?</h3>
        <p>No te preocupes, haz clic en el botón azul para crear una nueva:</p>
        <a href="{reset_url}" style="background-color: #007bff; color: white; padding: 12px 20px; text-decoration: none; border-radius: 5px; display: inline-block;">Restablecer mi Contraseña</a>
        <p>Si el botón no funciona, copia este link en tu navegador: <br> {reset_url}</p>
        <p>Este enlace dejará de funcionar en 1 hora por tu seguridad.</p>
    </body></html>
    """
    message.attach(MIMEText(text, "plain"))
    message.attach(MIMEText(html, "html"))

    try:
        context = ssl.create_default_context()
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls(context=context) # Encriptamos la comunicación para seguridad
            server.login(smtp_user, smtp_password)
            server.sendmail(sender_email, to_email, message.as_string())
        return True
    except Exception as e:
        print(f"Error al intentar enviar el correo de recuperación: {e}")
        return False
