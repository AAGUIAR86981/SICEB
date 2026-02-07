from flask import Blueprint, render_template, session, flash, redirect, url_for, request, Response
from utils.decorators import login_required, permission_required
from datetime import datetime
import logging
import csv
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from models.provision import Provision
from models.employee import Employee
from models.combos_model import ComboModel

# Configurar logging
logger = logging.getLogger(__name__)

provision_bp = Blueprint('provision', __name__)

# Se han eliminado las funciones auxiliares (get_tipo_provision_actual, get_tablas_provision, etc.)
# porque ahora residen en los modelos Provision y Employee.


# =============================================================================
# RUTA 1: CONFIGURAR PROVISIÓN (SOLO ADMIN Y SUPERVISOR)
# =============================================================================
@provision_bp.route('/configurar_provision', methods=['GET', 'POST'])
@login_required
@permission_required('create_provisions')  # ← AGREGADO
def configurar_provision():
    """Configurar una nueva provisión - Solo admin y supervisor"""
    try:
        if request.method == 'POST':
            semana_prov = request.form.get('SemanaProv')
            tipo_nomina = request.form.get('Nomina')
            fecha_inicial = request.form.get('fecha_inicial')

            # Validaciones básicas
            if not all([semana_prov, tipo_nomina, fecha_inicial]):
                flash('Todos los campos son obligatorios', 'error')
                return render_template('configurar_provision.html')

            # Aquí iría tu lógica para guardar la configuración
            # Por ahora solo redirigimos
            flash('Configuración de provisión guardada exitosamente', 'success')
            return redirect(url_for('provision.make_provision'))

        else:
            # MOSTRAR FORMULARIO DE CONFIGURACIÓN
            return render_template('configurar_provision.html')

    except Exception as e:
        logger.error(f"Error en configurar_provision: {e}")
        flash('Error al configurar la provisión', 'error')
        return render_template('configurar_provision.html')


# =============================================================================
# RUTA 2: REALIZAR PROVISIÓN (SOLO ADMIN Y SUPERVISOR)
# =============================================================================
@provision_bp.route('/provision', methods=['GET', 'POST'])
@login_required
@permission_required('create_provisions') 
def make_provision():
    """Maneja el formulario y procesamiento de provisiones - Solo admin y supervisor"""
    if request.method == 'POST':
        return procesar_provision_post()
    else:
        return mostrar_formulario_provision_get()


def mostrar_formulario_provision_get():
    """Muestra el formulario para realizar provisión (GET)"""
    try:
        # Obtener CONFIGURACIONES disponibles
        config_semanal, config_quincenal = Provision.get_available_tables()
        
        # Determinar periodos actuales
        semana_iso, quincena = Provision.get_type_and_week()
        
        # Por defecto, sugerimos el tipo basado en el día (1-15: Quincenal, 16+: Semanal)
        # pero permitimos que el usuario elija.
        dia_actual = datetime.now().day
        tipo_sugerido = 'quincenal' if dia_actual <= 15 else 'semanal'
        
        config = config_quincenal if tipo_sugerido == 'quincenal' else config_semanal
        semProv = quincena if tipo_sugerido == 'quincenal' else semana_iso

        if not config:
             # Si el sugerido no existe, probamos el otro
             tipo_sugerido = 'semanal' if tipo_sugerido == 'quincenal' else 'quincenal'
             config = config_semanal if tipo_sugerido == 'semanal' else config_quincenal
             semProv = semana_iso if tipo_sugerido == 'semanal' else quincena

        if not config:
            flash('Error: No se encontraron configuraciones de provisión activas', 'error')
            combos = Provision.get_active_combos()
            return render_template('realizarP.html', tipo_provision='semanal', combos=combos)

        # Obtener productos de la provisión
        productos = Provision.get_products(config['id'], semProv)

        # Preparar datos para la sesión y template
        session['combo'] = productos
        session['fecha'] = datetime.now().strftime('%A, %d / %m / %Y')
        session['semana'] = semProv
        session['tipo_provision'] = tipo_sugerido
        
        # Guardar config_id y periodos en sesión para el POST
        session['provision_config_id'] = config['id']
        session['semana_iso'] = semana_iso
        session['quincena'] = quincena

        # Obtener combos disponibles
        combos = Provision.get_active_combos()

        logger.info(f"Formulario provision - Tipo Sugerido: {tipo_sugerido}, Semana/Periodo: {semProv}")
        return render_template('realizarP.html', tipo_provision=tipo_sugerido, combos=combos)

    except Exception as e:
        logger.error(f"Error en mostrar_formulario_provision_get: {e}")
        flash('Error al cargar el formulario de provisión', 'error')
        return redirect(url_for('auth.index'))


def procesar_provision_post():
    """Procesa la provisión y muestra resultados (POST)"""
    try:
        # 1. Validar datos del formulario
        tipo_nomina = request.form.get('Nomina') # '1' (Semanal), '2' (Quincenal)

        if not tipo_nomina:
            flash('Debe seleccionar el tipo de nómina', 'error')
            return redirect(url_for('provision.make_provision'))

        # 2. Determinar periodo exacto basado en la selección del usuario
        semana_iso, quincena = Provision.get_type_and_week()
        
        # El periodo (semana base) depende de lo que el usuario seleccionó
        periodo_actual = semana_iso if tipo_nomina == '1' else quincena
        
        # 3. Verificar duplicados
        if Provision.exists(periodo_actual, tipo_nomina):
            nomina_lbl = 'Semanal' if tipo_nomina == '1' else 'Quincenal'
            flash(f'Error: Ya se ha generado la provisión de nómina {nomina_lbl} para el día de hoy. Solo se permite una por día.', 'error')
            return redirect(url_for('provision.make_provision'))
        
        # Usamos la config de la sesión si existe, sino buscamos
        config_id = session.get('provision_config_id')
        tipo_prov_slug = 'semanal' if tipo_nomina == '1' else 'quincenal'
        
        if not config_id:
             config_semanal, config_quincenal = Provision.get_available_tables()
             config = config_semanal if tipo_prov_slug == 'semanal' else config_quincenal
             if config:
                 config_id = config['id']
        
        if not config_id:
            flash('Error de configuración de provisión', 'error')
            return redirect(url_for('provision.make_provision'))
        
        # Necesitamos semProv para el log y guardado
        # Para semanal: semana_iso. Para quincenal: quincena.
        semProv = periodo_actual
        tipo_provision = tipo_prov_slug

        # 4. Obtener productos de la provisión (ya sea combo o estándar)
        combo_id = request.form.get('combo_id')
        if combo_id and combo_id != 'standard':
            selected_combo = ComboModel.get_combo_by_id(int(combo_id))
            if selected_combo:
                productos = [(item['nombre'], str(item['cantidad'])) for item in selected_combo['items']]
            else:
                productos = Provision.get_products(config_id, semProv)
        else:
            productos = Provision.get_products(config_id, semProv)

        # 5. Obtener empleados
        asignados = Employee.get_all(tipo_nomina, 'activo', 1000)
        
        # Filtrar invalidados: Solo admin o roles con permiso especial pueden verlos
        user_permissions = session.get('user_permissions', [])
        can_view_unapproved = 'all' in user_permissions or 'view_unapproved' in user_permissions or session.get('isAdmin') == 1
        
        invalidados = []
        if can_view_unapproved:
            invalidados = Employee.get_all(tipo_nomina, 'inactivo', 1000)
        
        # Si no pueden ver invalidados, no deberíamos intentar guardarlos en logs con conteo real 
        # (o sí, para integridad del sistema, pero el usuario no los ve).
        # Mantendremos la integridad del sistema guardando el log correcto, pero ocultándolos en la vista.

        # 6. Guardar en el log de provisiones
        Provision.save_log(semProv, tipo_nomina, session.get('user', 'Usuario desconocido'), len(asignados), len(invalidados))

        # 7. Guardar en el historial de provisiones con información completa
        success = False
        if productos and (asignados or invalidados):
            success = Provision.save_history(
                tipo_provision,
                semProv,
                tipo_nomina,
                productos,
                asignados,
                invalidados,
                session.get('id'),
                session.get('user', 'Usuario desconocido')
            )

        if not success:
            logger.error(f"Error al guardar historial de provision - Tipo: {tipo_provision}, Periodo: {semProv}")
            flash('Error técnico: La provisión se procesó pero no pudo guardarse en el historial. Por favor contacte a soporte.', 'error')
            # Podríamos redirigir, pero dejamos que vea los resultados actuales en sesión por ahora

        # 8. Preparar datos para la plantilla
        session.update({
            'asignados': asignados,
            'invalidados': invalidados,
            'nomina': tipo_nomina,
            'semana': semProv,
            'semProv': semProv,
            'tipo_provision': tipo_provision,
            'fecha': datetime.now().strftime('%A, %d / %m / %Y'),
            'combo': productos,
            'provExiste': False
        })

        logger.info(f"Provision procesada - Asignados: {len(asignados)}, Invalidados: {len(invalidados)}")
        flash('Provisión creada', 'success')
        return render_template('resultados_provision.html')

    except Exception as e:
        logger.error(f"Error crítico en procesar_provision_post: {e}")
        flash('Error al procesar la provisión', 'error')
        return redirect(url_for('provision.make_provision'))


# =============================================================================
# RUTAS DE EXPORTACIÓN (SOLO ADMIN Y SUPERVISOR)
# =============================================================================
@provision_bp.route('/exportar/<tipo_empleado>/<formato>')
@login_required
@permission_required('create_provisions')  # ← AGREGADO
def exportar_empleados(tipo_empleado, formato):
    """Exporta empleados a Excel, CSV - Solo admin y supervisor"""
    try:
        if tipo_empleado == 'asignados':
            datos = session.get('asignados', [])
            filename = f"empleados_asignados_{datetime.now().strftime('%Y%m%d_%H%M')}"
            titulo = "EMPLEADOS ASIGNADOS - PROVISIÓN"
        elif tipo_empleado == 'invalidados':
            # Verificar permiso para exportar invalidados
            user_permissions = session.get('user_permissions', [])
            can_view_unapproved = 'all' in user_permissions or 'view_unapproved' in user_permissions or session.get('isAdmin') == 1
            
            if not can_view_unapproved:
                flash('No tiene permisos para exportar empleados no aprobados', 'error')
                return redirect(url_for('provision.make_provision'))

            datos = session.get('invalidados', [])
            filename = f"empleados_invalidados_{datetime.now().strftime('%Y%m%d_%H%M')}"
            titulo = "EMPLEADOS INVALIDADOS - PROVISIÓN"
        else:
            flash('Tipo de empleado no válido', 'error')
            return redirect(url_for('provision.make_provision'))

        if not datos:
            flash('No hay datos para exportar', 'warning')
            return redirect(url_for('provision.make_provision'))

        # Encabezados de columnas
        headers = ['Cédula', 'Nombre', 'Apellido', 'Departamento', 'ID Empleado', 'Estatus']

        if formato == 'csv':
            return exportar_csv(datos, headers, filename, titulo)
        elif formato in ['xlsx', 'xls']:
            return exportar_excel(datos, headers, filename, titulo, formato)
        else:
            flash('Formato no soportado', 'error')
            return redirect(url_for('provision.make_provision'))

    except Exception as e:
        logger.error(f"Error en exportación: {e}")
        flash('Error al exportar los datos', 'error')
        return redirect(url_for('provision.make_provision'))


def exportar_csv(datos, headers, filename, titulo):
    """Exporta datos a formato CSV"""
    output = io.StringIO()
    writer = csv.writer(output)

    # Escribir título y encabezados
    writer.writerow([titulo])
    writer.writerow([f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    writer.writerow([])
    writer.writerow(headers)

    # Escribir datos
    for empleado in datos:
        estatus = "ASIGNADO" if "asignados" in filename else "INVALIDADO"
        writer.writerow(list(empleado) + [estatus])

    response = Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename={filename}.csv"}
    )

    return response


def exportar_excel(datos, headers, filename, titulo, formato):
    """Exporta datos a formato Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Empleados"

    # Estilos
    titulo_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Escribir título
    ws.merge_cells('A1:F1')
    ws['A1'] = titulo
    ws['A1'].font = titulo_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Escribir fecha
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal='center')

    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Escribir datos
    for row, empleado in enumerate(datos, 5):
        estatus = "ASIGNADO" if "asignados" in filename else "INVALIDADO"
        for col, valor in enumerate(list(empleado) + [estatus], 1):
            ws.cell(row=row, column=col, value=valor)

    # Ajustar anchos de columna
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    extension = "xlsx"

    response = Response(
        buffer.read(),
        mimetype=mimetype,
        headers={"Content-Disposition": f"attachment;filename={filename}.{extension}"}
    )

    return response


# =============================================================================
# RUTA PARA VER HISTORIAL DE PROVISIONES (TODOS LOS USUARIOS)
# =============================================================================
@provision_bp.route('/historial_provisiones')
@login_required  # ← Solo login_required, todos pueden ver histórico
def historial_provisiones():
    """Muestra el historial completo de provisiones realizadas - TODOS los usuarios autenticados"""
    print("DEBUG: Accessing historial_provisiones route") # DEBUG
    try:
        # Fetch raw data from model
        raw_history = Provision.get_history(100)
        print(f"DEBUG: raw_history count: {len(raw_history)}") # DEBUG

        historial = []
        for row in raw_history:
            try:
                # row structure: 0:id, 1:tipo_provision, 2:semana, 3:tipo_nomina, 4:productos, 
                # 5:datos_completos, 6:fecha, 7:usuario
                
                # Manejar productos
                productos = []
                if row[4]:  # productos
                    try:
                        productos = json.loads(row[4])
                    except:
                        productos = []

                # Manejar datos completos con valores por defecto
                datos_completos = {
                    'empleados_asignados': 0,
                    'empleados_invalidados': 0,
                    'resumen_nomina': {
                        'total_empleados': 0,
                        'empleados_activos': 0,
                        'empleados_inactivos': 0,
                        'total_departamentos': 0
                    },
                    'departamentos_detallados': []
                }

                if row[5]:  # datos_completos
                    try:
                        datos_cargados = json.loads(row[5])
                        if isinstance(datos_cargados, dict):
                            datos_completos.update(datos_cargados)
                            if 'resumen_nomina' not in datos_completos:
                                datos_completos['resumen_nomina'] = {
                                    'total_empleados': 0,
                                    'empleados_activos': 0,
                                    'empleados_inactivos': 0,
                                    'total_departamentos': 0
                                }
                    except:
                        pass

                historial.append({
                    'id': row[0],
                    'tipo_provision': row[1] or 'N/A',
                    'semana': row[2] or 'N/A',
                    'tipo_nomina': row[3] or 'N/A',
                    'productos': productos,
                    'datos_completos': datos_completos,
                    'fecha_creacion': row[6],
                    'usuario_nombre': row[7] or 'Usuario desconocido'
                })

            except Exception as e:
                import traceback
                traceback.print_exc()
                print(f"DEBUG ERROR parsing row: {e}")
                logger.error(f"Error procesando fila del historial: {e}")
                continue

        logger.info(f"Historial cargado: {len(historial)} registros")

        # Pasar información de rol al template
        es_admin = session.get('isAdmin') == 1
        user_roles = session.get('user_roles', [])
        es_supervisor = any(role.get('name') == 'supervisor' for role in user_roles)

        return render_template('historial_provisiones.html',
                             historial=historial,
                             es_admin=es_admin,
                             es_supervisor=es_supervisor)

    except Exception as e:
        logger.error(f"Error obteniendo historial: {e}")
        flash('Error al cargar el historial de provisiones', 'error')
        return render_template('historial_provisiones.html', historial=[])
