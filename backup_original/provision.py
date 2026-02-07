from flask import Blueprint, render_template, session, flash, redirect, url_for, request, Response
from config.database import get_db_connection
from utils.decorators import login_required
from datetime import datetime
import logging
import csv
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Configurar logging
logger = logging.getLogger(__name__)

provision_bp = Blueprint('provision', __name__)


def get_tipo_provision_actual():
    """Determina el tipo de provisión basado en la fecha actual"""
    serverDatetime = datetime.now()
    dia_del_mes = serverDatetime.day

    if dia_del_mes <= 15:
        return 'quincenal', 1
    else:
        return 'semanal', 2


def get_tablas_provision(cursor):
    """Obtiene las tablas de provisión disponibles"""
    cursor.execute("SHOW TABLES")
    tablas_existentes = [tabla[0] for tabla in cursor.fetchall()]

    tabla_semanal = 'semana_provision' if 'semana_provision' in tablas_existentes else None
    tabla_quincenal = 'semana_provision_quincenal' if 'semana_provision_quincenal' in tablas_existentes else None

    return tabla_semanal, tabla_quincenal


def obtener_productos_provision(cursor, tabla, semProv):
    """Obtiene los productos de la provisión configurada"""
    if not tabla:
        return []

    try:
        cursor.execute(f'SELECT * FROM {tabla} WHERE id=%s', (semProv,))
        listaProductos = cursor.fetchone()

        if not listaProductos:
            return []

        # Obtener estructura de columnas
        cursor.execute(f"SHOW COLUMNS FROM {tabla}")
        columnas = [col[0] for col in cursor.fetchall()]

        # DETECTAR automáticamente las columnas de productos y cantidades
        columnas_productos = []
        columnas_cantidades = []

        for col in columnas:
            col_lower = col.lower()
            if any(keyword in col_lower for keyword in ['rubro', 'producto']):
                columnas_productos.append(col)
            elif any(keyword in col_lower for keyword in ['cant', 'cantidad']):
                columnas_cantidades.append(col)

        # Ordenar las columnas para que coincidan
        columnas_productos.sort()
        columnas_cantidades.sort()

        prodCant = []
        for i in range(min(len(columnas_productos), len(columnas_cantidades))):
            rubro_idx = columnas.index(columnas_productos[i])
            cant_idx = columnas.index(columnas_cantidades[i])

            rubro = listaProductos[rubro_idx] if rubro_idx < len(listaProductos) else None
            cant = listaProductos[cant_idx] if cant_idx < len(listaProductos) else None

            if rubro and cant is not None:
                prodCant.append((rubro, str(cant)))

        logger.info(f"Productos obtenidos: {prodCant} desde tabla {tabla}")
        return prodCant

    except Exception as e:
        logger.error(f"Error obteniendo productos de provisión: {e}")
        return []


def obtener_empleados(cursor, tipo_nomina, estado='activo', limite=100):
    """Obtiene empleados de la base de datos"""
    try:
        # Convertir 'activo'/'inactivo' a 1/0 para boolValidacion
        bool_val = 1 if estado == 'activo' else 0

        query = '''
            SELECT cedula, nombre, apellido, departamento, id_empleado
            FROM empleados
            WHERE tipoNomina = %s AND boolValidacion = %s
            LIMIT %s
        '''
        cursor.execute(query, (int(tipo_nomina), bool_val, limite))
        empleados = cursor.fetchall()
        logger.info(f"Empleados obtenidos: {len(empleados)} para tipo_nomina={tipo_nomina}, estado={estado}")
        return empleados
    except Exception as e:
        logger.error(f"Error obteniendo empleados ({estado}): {e}")
        return []


def obtener_departamentos_detallados(cursor, tipo_nomina):
    """Obtiene departamentos con conteo de empleados por tipo de nómina"""
    try:
        query = '''
            SELECT departamento,
                   COUNT(CASE WHEN boolValidacion = 1 THEN 1 END) as activos,
                   COUNT(CASE WHEN boolValidacion = 0 THEN 1 END) as inactivos
            FROM empleados
            WHERE tipoNomina = %s
            GROUP BY departamento
            ORDER BY departamento
        '''
        cursor.execute(query, (int(tipo_nomina),))
        return cursor.fetchall()
    except Exception as e:
        logger.error(f"Error obteniendo departamentos: {e}")
        return []


def obtener_resumen_nomina(cursor, tipo_nomina):
    """Obtiene resumen completo por tipo de nómina"""
    try:
        query = '''
            SELECT
                COUNT(*) as total_empleados,
                COUNT(CASE WHEN boolValidacion = 1 THEN 1 END) as empleados_activos,
                COUNT(CASE WHEN boolValidacion = 0 THEN 1 END) as empleados_inactivos,
                COUNT(DISTINCT departamento) as total_departamentos
            FROM empleados
            WHERE tipoNomina = %s
        '''
        cursor.execute(query, (int(tipo_nomina),))
        return cursor.fetchone()
    except Exception as e:
        logger.error(f"Error obteniendo resumen nómina: {e}")
        return (0, 0, 0, 0)


def guardar_en_historial(cursor, tipo_provision, semana, tipo_nomina, productos, asignados, invalidados, usuario_id, usuario_nombre):
    """Guarda la provisión en el historial con información completa de otras tablas"""
    try:
        # Obtener información detallada de otras tablas
        departamentos_detallados = obtener_departamentos_detallados(cursor, tipo_nomina)
        resumen_nomina = obtener_resumen_nomina(cursor, tipo_nomina)

        # Preparar datos para el historial
        datos_historial = {
            'productos': productos,
            'tipo_nomina': 'Semanal' if tipo_nomina == '1' else 'Quincenal',
            'empleados_asignados': len(asignados),
            'empleados_invalidados': len(invalidados),
            'resumen_nomina': {
                'total_empleados': resumen_nomina[0] if resumen_nomina else 0,
                'empleados_activos': resumen_nomina[1] if resumen_nomina else 0,
                'empleados_inactivos': resumen_nomina[2] if resumen_nomina else 0,
                'total_departamentos': resumen_nomina[3] if resumen_nomina else 0
            },
            'departamentos_detallados': [
                {
                    'departamento': depto[0],
                    'activos': depto[1],
                    'inactivos': depto[2]
                } for depto in departamentos_detallados
            ],
            'fecha_procesamiento': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'usuario': usuario_nombre
        }

        productos_json = json.dumps(productos, ensure_ascii=False)
        datos_completos_json = json.dumps(datos_historial, ensure_ascii=False)

        cursor.execute('''
            INSERT INTO provisiones_historial
            (tipo_provision, semana, tipo_nomina, productos, datos_completos, usuario_id, usuario_nombre)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        ''', (
            tipo_provision,
            semana,
            'Semanal' if tipo_nomina == '1' else 'Quincenal',
            productos_json,
            datos_completos_json,
            usuario_id,
            usuario_nombre
        ))

        logger.info(f"Provision guardada en historial - Tipo: {tipo_provision}, Semana: {semana}, Nómina: {tipo_nomina}")
        return True
    except Exception as e:
        logger.error(f"Error guardando en historial: {e}")
        return False


# =============================================================================
# RUTA 1: CONFIGURAR PROVISIÓN
# =============================================================================
@provision_bp.route('/configurar_provision', methods=['GET', 'POST'])
@login_required
def configurar_provision():
    """Configurar una nueva provisión - Formulario con pasos"""
    try:
        if request.method == 'POST':
            semana_prov = request.form.get('SemanaProv')
            tipo_nomina = request.form.get('Nomina')
            fecha_inicial = request.form.get('fecha_inicial')

            # Validaciones básicas
            if not all([semana_prov, tipo_nomina, fecha_inicial]):
                flash('Todos los campos son obligatorios', 'error')
                return render_template('configurar_provision.html')

            # Aquí iría tu lógica para guardar la configuración
            # Por ahora solo redirigimos
            flash('Configuración de provisión guardada exitosamente', 'success')
            return redirect(url_for('provision.make_provision'))

        else:
            # MOSTRAR FORMULARIO DE CONFIGURACIÓN
            return render_template('configurar_provision.html')

    except Exception as e:
        logger.error(f"Error en configurar_provision: {e}")
        flash('Error al configurar la provisión', 'error')
        return render_template('configurar_provision.html')


# =============================================================================
# RUTA 2: REALIZAR PROVISIÓN (PRINCIPAL)
# =============================================================================
@provision_bp.route('/provision', methods=['GET', 'POST'])
@login_required
def make_provision():
    """Maneja el formulario y procesamiento de provisiones"""
    if request.method == 'POST':
        return procesar_provision_post()
    else:
        return mostrar_formulario_provision_get()


def mostrar_formulario_provision_get():
    """Muestra el formulario para realizar provisión (GET)"""
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Obtener tablas disponibles
        tabla_semanal, tabla_quincenal = get_tablas_provision(cursor)

        if not tabla_semanal and not tabla_quincenal:
            flash('Error: No se encontraron tablas de provisión en la base de datos', 'error')
            return render_template('realizarP.html', tipo_provision='semanal')

        # Determinar tipo de provisión actual
        tipo_provision, semProv = get_tipo_provision_actual()
        tabla = tabla_quincenal if tipo_provision == 'quincenal' else tabla_semanal

        if tabla not in [tabla_semanal, tabla_quincenal]:
            flash(f'Error: La tabla {tabla} no existe', 'error')
            return render_template('realizarP.html', tipo_provision=tipo_provision)

        # Obtener productos de la provisión
        productos = obtener_productos_provision(cursor, tabla, semProv)

        # Preparar datos para la sesión y template
        session['combo'] = productos
        session['fecha'] = datetime.now().strftime('%A, %d / %m / %Y')
        session['semProv'] = semProv
        session['tipo_provision'] = tipo_provision

        if not productos:
            flash('No se encontró provisión configurada para este período', 'warning')

        logger.info(f"Formulario provision mostrado - Tipo: {tipo_provision}, Productos: {len(productos)}")
        return render_template('realizarP.html', tipo_provision=tipo_provision)

    except Exception as e:
        logger.error(f"Error en mostrar_formulario_provision_get: {e}")
        flash('Error al cargar el formulario de provisión', 'error')
        return redirect(url_for('auth.index'))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def procesar_provision_post():
    """Procesa la provisión y muestra resultados (POST)"""
    conn = None
    cursor = None
    try:
        # 1. Validar datos del formulario
        tipo_nomina = request.form.get('Nomina')

        if not tipo_nomina:
            flash('Debe seleccionar el tipo de nómina', 'error')
            return redirect(url_for('provision.make_provision'))

        # 2. Conectar a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor()

        # 3. Determinar configuración de provisión
        tipo_provision, semProv = get_tipo_provision_actual()
        tabla_semanal, tabla_quincenal = get_tablas_provision(cursor)
        tabla = tabla_quincenal if tipo_provision == 'quincenal' else tabla_semanal

        # 4. Obtener productos de la provisión
        productos = obtener_productos_provision(cursor, tabla, semProv)

        # 5. Obtener empleados
        asignados = obtener_empleados(cursor, tipo_nomina, 'activo', 1000)
        invalidados = obtener_empleados(cursor, tipo_nomina, 'inactivo', 1000)

        # 6. Guardar en el log de provisiones
        try:
            cursor.execute('''
                INSERT INTO prov_logs
                (semana, tipoNom, Usuario, fechaProv, cantAprob, cantRecha)
                VALUES (%s, %s, %s, NOW(), %s, %s)
            ''', (
                str(semProv),
                'Semanal' if tipo_nomina == '1' else 'Quincenal',
                session.get('user', 'Usuario desconocido'),
                len(asignados),
                len(invalidados)
            ))
            conn.commit()
            logger.info(f"Provision guardada en logs - Aprobados: {len(asignados)}, Rechazados: {len(invalidados)}")
        except Exception as e:
            logger.warning(f"No se pudo guardar en prov_logs: {e}")
            conn.rollback()

        # 7. Guardar en el historial de provisiones con información completa
        if productos and (asignados or invalidados):
            guardar_en_historial(
                cursor,
                tipo_provision,
                semProv,
                tipo_nomina,
                productos,
                asignados,
                invalidados,
                session.get('id'),
                session.get('user', 'Usuario desconocido')
            )
            conn.commit()

        # 8. Preparar datos para la plantilla
        session.update({
            'asignados': asignados,
            'invalidados': invalidados,
            'nomina': tipo_nomina,
            'semana': semProv,
            'semProv': semProv,
            'tipo_provision': tipo_provision,
            'fecha': datetime.now().strftime('%A, %d / %m / %Y'),
            'combo': productos,
            'provExiste': False
        })

        logger.info(f"Provision procesada - Asignados: {len(asignados)}, Invalidados: {len(invalidados)}")
        flash('Provisión procesada exitosamente', 'success')
        return render_template('resultados_provision.html')

    except Exception as e:
        logger.error(f"Error crítico en procesar_provision_post: {e}")
        if conn:
            conn.rollback()
        flash('Error al procesar la provisión', 'error')
        return redirect(url_for('provision.make_provision'))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


# =============================================================================
# RUTAS DE EXPORTACIÓN
# =============================================================================
@provision_bp.route('/exportar/<tipo_empleado>/<formato>')
@login_required
def exportar_empleados(tipo_empleado, formato):
    """Exporta empleados a Excel, CSV u otros formatos"""
    try:
        if tipo_empleado == 'asignados':
            datos = session.get('asignados', [])
            filename = f"empleados_asignados_{datetime.now().strftime('%Y%m%d_%H%M')}"
            titulo = "EMPLEADOS ASIGNADOS - PROVISIÓN"
        elif tipo_empleado == 'invalidados':
            datos = session.get('invalidados', [])
            filename = f"empleados_invalidados_{datetime.now().strftime('%Y%m%d_%H%M')}"
            titulo = "EMPLEADOS INVALIDADOS - PROVISIÓN"
        else:
            flash('Tipo de empleado no válido', 'error')
            return redirect(url_for('provision.make_provision'))

        if not datos:
            flash('No hay datos para exportar', 'warning')
            return redirect(url_for('provision.make_provision'))

        # Encabezados de columnas
        headers = ['Cédula', 'Nombre', 'Apellido', 'Departamento', 'ID Empleado', 'Estatus']

        if formato == 'csv':
            return exportar_csv(datos, headers, filename, titulo)
        elif formato in ['xlsx', 'xls']:
            return exportar_excel(datos, headers, filename, titulo, formato)
        else:
            flash('Formato no soportado', 'error')
            return redirect(url_for('provision.make_provision'))

    except Exception as e:
        logger.error(f"Error en exportación: {e}")
        flash('Error al exportar los datos', 'error')
        return redirect(url_for('provision.make_provision'))


def exportar_csv(datos, headers, filename, titulo):
    """Exporta datos a formato CSV"""
    output = io.StringIO()
    writer = csv.writer(output)

    # Escribir título y encabezados
    writer.writerow([titulo])
    writer.writerow([f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    writer.writerow([])
    writer.writerow(headers)

    # Escribir datos
    for empleado in datos:
        estatus = "ASIGNADO" if "asignados" in filename else "INVALIDADO"
        writer.writerow(list(empleado) + [estatus])

    response = Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename={filename}.csv"}
    )

    return response


def exportar_excel(datos, headers, filename, titulo, formato):
    """Exporta datos a formato Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Empleados"

    # Estilos
    titulo_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Escribir título
    ws.merge_cells('A1:F1')
    ws['A1'] = titulo
    ws['A1'].font = titulo_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Escribir fecha
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal='center')

    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Escribir datos
    for row, empleado in enumerate(datos, 5):
        estatus = "ASIGNADO" if "asignados" in filename else "INVALIDADO"
        for col, valor in enumerate(list(empleado) + [estatus], 1):
            ws.cell(row=row, column=col, value=valor)

    # Ajustar anchos de columna
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    extension = "xlsx"

    response = Response(
        buffer.read(),
        mimetype=mimetype,
        headers={"Content-Disposition": f"attachment;filename={filename}.{extension}"}
    )

    return response


# =============================================================================
# RUTA PARA VER HISTORIAL DE PROVISIONES
# =============================================================================
@provision_bp.route('/historial_provisiones')
@login_required
def historial_provisiones():
    """Muestra el historial completo de provisiones realizadas"""
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT id, tipo_provision, semana, tipo_nomina, productos, datos_completos, 
                   fecha_creacion, usuario_nombre
            FROM provisiones_historial
            ORDER BY fecha_creacion DESC
            LIMIT 100
        ''')

        historial = []
        for row in cursor.fetchall():
            try:
                # Manejar productos
                productos = []
                if row[4]:  # productos
                    try:
                        productos = json.loads(row[4])
                    except:
                        productos = []

                # Manejar datos completos con valores por defecto
                datos_completos = {
                    'empleados_asignados': 0,
                    'empleados_invalidados': 0,
                    'resumen_nomina': {
                        'total_empleados': 0,
                        'empleados_activos': 0,
                        'empleados_inactivos': 0,
                        'total_departamentos': 0
                    },
                    'departamentos_detallados': []
                }

                if row[5]:  # datos_completos
                    try:
                        datos_cargados = json.loads(row[5])
                        # Actualizar solo los campos que existen
                        if isinstance(datos_cargados, dict):
                            datos_completos.update(datos_cargados)
                            # Asegurar que resumen_nomina existe
                            if 'resumen_nomina' not in datos_completos:
                                datos_completos['resumen_nomina'] = {
                                    'total_empleados': 0,
                                    'empleados_activos': 0,
                                    'empleados_inactivos': 0,
                                    'total_departamentos': 0
                                }
                    except:
                        # Si hay error, mantener los valores por defecto
                        pass

                historial.append({
                    'id': row[0],
                    'tipo_provision': row[1] or 'N/A',
                    'semana': row[2] or 'N/A',
                    'tipo_nomina': row[3] or 'N/A',
                    'productos': productos,
                    'datos_completos': datos_completos,
                    'fecha_creacion': row[6],
                    'usuario_nombre': row[7] or 'Usuario desconocido'
                })

            except Exception as e:
                logger.error(f"Error procesando fila del historial: {e}")
                continue

        logger.info(f"Historial cargado: {len(historial)} registros")
        return render_template('historial_provisiones.html', historial=historial)

    except Exception as e:
        logger.error(f"Error obteniendo historial: {e}")
        flash('Error al cargar el historial de provisiones', 'error')
        return render_template('historial_provisiones.html', historial=[])
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
