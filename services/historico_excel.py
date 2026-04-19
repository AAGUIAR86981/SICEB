# services/historico_excel.py
from flask import send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO
import os
from PIL import Image as PILImage
from datetime import datetime

def generar_reporte_historico_excel(logs, total_records, totalAprob, totalRecha, total):
    """
    Genera un archivo Excel personalizado con el historial de provisiones
    """
    wb = openpyxl.Workbook()
    
    # Hoja 1: Resumen Estadístico
    ws_resumen = wb.active
    ws_resumen.title = "Resumen Estadístico"
    
    # Hoja 2: Detalle de Provisiones
    ws_detalle = wb.create_sheet("Detalle de Provisiones")
    
    # ============ ESTILOS ============
    titulo_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    titulo_fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
    titulo_alignment = Alignment(horizontal='center', vertical='center')
    
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0a58ca', end_color='0a58ca', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # BORDE - Agregado para todas las celdas
    borde = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    data_alignment_center = Alignment(horizontal='center', vertical='center')
    data_alignment_left = Alignment(horizontal='left', vertical='center')
    
    membrete_font = Font(name='Arial', size=8, bold=False, color='444444')
    membrete_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # ============ AGREGAR LOGO ============
    logo_encontrado = None
    formatos = ['png', 'PNG', 'webp', 'WEBP', 'jpg', 'JPG', 'jpeg', 'JPEG']
    
    for formato in formatos:
        ruta_logo = os.path.join('static', 'images', f'imagen1.{formato}')
        if os.path.exists(ruta_logo):
            logo_encontrado = ruta_logo
            break
    
    if not logo_encontrado:
        import glob
        archivos = glob.glob(os.path.join('static', 'images', 'imagen1.*'))
        if archivos:
            logo_encontrado = archivos[0]
    
    def agregar_membrete(ws, columnas='H'):
        """Agrega el membrete a una hoja"""
        if logo_encontrado and os.path.exists(logo_encontrado):
            try:
                if logo_encontrado.endswith('.webp'):
                    img_pil = PILImage.open(logo_encontrado)
                    temp_png = BytesIO()
                    img_pil.save(temp_png, format='PNG')
                    temp_png.seek(0)
                    img = Image(temp_png)
                else:
                    img = Image(logo_encontrado)
                
                img.width = 60
                img.height = 60
                ws.add_image(img, 'A1')
            except Exception as e:
                print(f"Error al cargar el logo: {e}")
        
        ws.merge_cells(f'B1:{columnas}1')
        celda_membrete = ws['B1']
        celda_membrete.value = "L.P. LIDER POLLO, C.A. RIF J-30713528-0"
        celda_membrete.font = membrete_font
        celda_membrete.alignment = membrete_alignment
        
        ws.merge_cells(f'B2:{columnas}2')
        celda_direccion = ws['B2']
        celda_direccion.value = "Zona Industrial El Recreo Parcela Nro. 51 Vía Flor Amarillo, Valencia Edo. Carabobo"
        celda_direccion.font = membrete_font
        celda_direccion.alignment = membrete_alignment
        
        ws.merge_cells(f'B3:{columnas}3')
        celda_contacto = ws['B3']
        celda_contacto.value = "Telf: (0412) 801.7687 / 801.7887 - liderpolloca@gmail.com"
        celda_contacto.font = membrete_font
        celda_contacto.alignment = membrete_alignment
        
        ws.row_dimensions[4].height = 10
    
    # ============ HOJA DE RESUMEN ============
    agregar_membrete(ws_resumen, 'G')
    
    # Título
    ws_resumen.merge_cells('A5:G5')
    celda_titulo = ws_resumen['A5']
    celda_titulo.value = "REPORTE HISTÓRICO DE PROVISIONES - RESUMEN"
    celda_titulo.font = titulo_font
    celda_titulo.fill = titulo_fill
    celda_titulo.alignment = titulo_alignment
    ws_resumen.row_dimensions[5].height = 30
    
    # Subtítulo
    ws_resumen.merge_cells('A6:G6')
    celda_subtitulo = ws_resumen['A6']
    fecha_reporte = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    celda_subtitulo.value = f"Generado: {fecha_reporte}"
    celda_subtitulo.font = Font(size=9, italic=True, color='666666')
    celda_subtitulo.alignment = Alignment(horizontal='center')
    ws_resumen.row_dimensions[6].height = 20
    
    # Tarjetas de estadísticas CON BORDES
    stats = [
        ['Total de Provisiones', total_records, 'primary'],
        ['Total Aprobados', totalAprob, 'success'],
        ['Total Rechazados', totalRecha, 'danger'],
        ['Total General', total, 'info']
    ]
    
    for idx, stat in enumerate(stats, start=8):
        ws_resumen.merge_cells(f'A{idx}:C{idx}')
        celda_label = ws_resumen[f'A{idx}']
        celda_label.value = stat[0]
        celda_label.font = Font(bold=True, size=12)
        celda_label.fill = PatternFill(start_color='e7f3ff', end_color='e7f3ff', fill_type='solid')
        celda_label.alignment = Alignment(horizontal='center')
        #celda_label.border = borde  # BORDE AGREGADO
        
        ws_resumen.merge_cells(f'D{idx}:G{idx}')
        celda_valor = ws_resumen[f'D{idx}']
        celda_valor.value = stat[1]
        celda_valor.alignment = Alignment(horizontal='center')
        #celda_valor.border = borde  # BORDE AGREGADO
        
        if stat[2] == 'success':
            celda_valor.font = Font(bold=True, size=14, color='28a745')
        elif stat[2] == 'danger':
            celda_valor.font = Font(bold=True, size=14, color='dc3545')
        elif stat[2] == 'info':
            celda_valor.font = Font(bold=True, size=14, color='17a2b8')
    
    # Porcentajes
    ws_resumen.cell(row=14, column=1, value="ANÁLISIS PORCENTUAL").font = Font(bold=True, size=12)
    ws_resumen.merge_cells('A14:G14')
    ws_resumen['A14'].alignment = Alignment(horizontal='center')
    
    porcentaje_aprob = (totalAprob / total * 100) if total > 0 else 0
    porcentaje_recha = (totalRecha / total * 100) if total > 0 else 0
    
    ws_resumen.cell(row=15, column=1, value="Aprobados:").font = Font(bold=True)
    ws_resumen.cell(row=15, column=2, value=f"{porcentaje_aprob:.2f}%").font = Font(color='28a745', bold=True)
    
    ws_resumen.cell(row=16, column=1, value="Rechazados:").font = Font(bold=True)
    ws_resumen.cell(row=16, column=2, value=f"{porcentaje_recha:.2f}%").font = Font(color='dc3545', bold=True)
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 25
    ws_resumen.column_dimensions['D'].width = 15
    
    # ============ HOJA DE DETALLE ============
    agregar_membrete(ws_detalle, 'H')
    
    # Título
    ws_detalle.merge_cells('A5:H5')
    celda_titulo_detalle = ws_detalle['A5']
    celda_titulo_detalle.value = "REPORTE HISTÓRICO DE PROVISIONES - DETALLE"
    celda_titulo_detalle.font = titulo_font
    celda_titulo_detalle.fill = titulo_fill
    celda_titulo_detalle.alignment = titulo_alignment
    ws_detalle.row_dimensions[5].height = 30
    
    # Subtítulo
    ws_detalle.merge_cells('A6:H6')
    celda_sub_detalle = ws_detalle['A6']
    celda_sub_detalle.value = f"Generado: {fecha_reporte}"
    celda_sub_detalle.font = Font(size=9, italic=True, color='666666')
    celda_sub_detalle.alignment = Alignment(horizontal='center')
    ws_detalle.row_dimensions[6].height = 20
    
    # Encabezados CON BORDES
    headers = ['Semana', 'Tipo Nómina', 'Aprobados', 'Rechazados', 'Total', 'Usuario', 'IP', 'Fecha Creación']
    for col_num, header in enumerate(headers, 1):
        celda = ws_detalle.cell(row=8, column=col_num)
        celda.value = header
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = header_alignment
        celda.border = borde  # BORDE AGREGADO
    
    # Datos CON BORDES
    for row_num, log in enumerate(logs, start=9):
        tipo_nomina = 'Semanal' if str(log.get('tipo_nomina')) == '1' else 'Quincenal'
        total_fila = (log.get('cant_aprobados') or 0) + (log.get('cant_rechazados') or 0)
        fecha = log.get('fecha_creacion')
        fecha_str = fecha.strftime('%d/%m/%Y %H:%M') if fecha else ''
        
        # Semana
        celda = ws_detalle.cell(row=row_num, column=1, value=log.get('semana', ''))
        celda.alignment = data_alignment_center
        celda.border = borde  # BORDE AGREGADO
        
        # Tipo Nómina
        celda = ws_detalle.cell(row=row_num, column=2, value=tipo_nomina)
        celda.alignment = data_alignment_center
        celda.border = borde  # BORDE AGREGADO
        
        # Aprobados
        celda = ws_detalle.cell(row=row_num, column=3, value=log.get('cant_aprobados') or 0)
        celda.alignment = data_alignment_center
        celda.border = borde  # BORDE AGREGADO
        celda.font = Font(color='28a745', bold=True)
        
        # Rechazados
        celda = ws_detalle.cell(row=row_num, column=4, value=log.get('cant_rechazados') or 0)
        celda.alignment = data_alignment_center
        celda.border = borde  # BORDE AGREGADO
        celda.font = Font(color='dc3545', bold=True)
        
        # Total
        celda = ws_detalle.cell(row=row_num, column=5, value=total_fila)
        celda.alignment = data_alignment_center
        celda.border = borde  # BORDE AGREGADO
        
        # Usuario
        celda = ws_detalle.cell(row=row_num, column=6, value=log.get('usuario_nombre', ''))
        celda.alignment = data_alignment_left
        celda.border = borde  # BORDE AGREGADO
        
        # IP
        celda = ws_detalle.cell(row=row_num, column=7, value=log.get('ip_address') or '-')
        celda.alignment = data_alignment_left
        celda.border = borde  # BORDE AGREGADO
        
        # Fecha
        celda = ws_detalle.cell(row=row_num, column=8, value=fecha_str)
        celda.alignment = data_alignment_center
        celda.border = borde  # BORDE AGREGADO
    
    # Ajustar anchos
    anchos = {'A': 12, 'B': 18, 'C': 12, 'D': 12, 'E': 10, 'F': 25, 'G': 15, 'H': 20}
    for col, ancho in anchos.items():
        ws_detalle.column_dimensions[col].width = ancho
    
    # Total CON BORDE
    ultima_fila = len(logs) + 9
    ws_detalle.merge_cells(f'A{ultima_fila + 1}:H{ultima_fila + 1}')
    celda_total = ws_detalle.cell(row=ultima_fila + 1, column=1)
    celda_total.value = f"TOTAL DE REGISTROS: {len(logs)}"
    celda_total.font = Font(bold=True, size=11)
    celda_total.alignment = Alignment(horizontal='center')
    celda_total.fill = PatternFill(start_color='e7f3ff', end_color='e7f3ff', fill_type='solid')
    celda_total.border = borde  # BORDE AGREGADO
    
    # ============ GUARDAR ============
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'historico_provisiones_{timestamp}.xlsx'
    
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)