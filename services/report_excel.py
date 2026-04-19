# services/report_excel.py
from flask import send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime
import os

# ============ FUNCIÓN PARA REPORTE DE EMPLEADOS ============
def generar_reporte_empleados_excel(empleados):
    """
    Genera un archivo Excel personalizado con la lista de empleados.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Empleados"

    # Estilos
    titulo_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    titulo_fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
    titulo_alignment = Alignment(horizontal='center', vertical='center')
    titulo_alignment = Alignment(horizontal='center', vertical='center')

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0a58ca', end_color='0a58ca', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    data_alignment = Alignment(horizontal='left', vertical='center')
    data_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    membrete_font = Font(name='Arial', size=8, bold=False, color='444444')
    membrete_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ============ LOGO ============
    ruta_logo = os.path.join('static', 'images', 'Imagen1.jpg')
    if os.path.exists(ruta_logo):
        try:
            img = Image(ruta_logo)
            img.width = 60
            img.height = 60
            ws.add_image(img, 'A1')
        except:
            pass
    
    # ============ MEMBRETE ============
    ws.merge_cells('B1:F1')
    ws['B1'].value = "L.P. LIDER POLLO, C.A. RIF J-30713528-0"
    ws['B1'].font = membrete_font
    ws['B1'].alignment = membrete_alignment
    
    ws.merge_cells('B2:F2')
    ws['B2'].value = "Zona Industrial El Recreo Parcela Nro. 51 Vía Flor Amarillo, Valencia Edo. Carabobo"
    ws['B2'].font = membrete_font
    ws['B2'].alignment = membrete_alignment
    
    ws.merge_cells('B3:F3')
    ws['B3'].value = "Telf: (0412) 801.7687 / 801.7887 - liderpolloca@gmail.com"
    ws['B3'].font = membrete_font
    ws['B3'].alignment = membrete_alignment
    
    ws.row_dimensions[4].height = 10

    # Título
    ws.merge_cells('A5:F5')
    ws['A5'].value = "REPORTE DE EMPLEADOS - SICEB"
    ws['A5'].font = titulo_font
    ws['A5'].fill = titulo_fill
    ws['A5'].alignment = titulo_alignment
    ws.row_dimensions[5].height = 30

    # Fecha
    ws.merge_cells('A6:F6')
    ws['A6'].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A6'].font = Font(size=8)
    ws['A6'].alignment = Alignment(horizontal='center')

    # Encabezados
    headers = ['ID', 'Cédula', 'Nombre', 'Apellido', 'Departamento', 'Estado']
    for col_num, header in enumerate(headers, 1):
        celda = ws.cell(row=8, column=col_num)
        celda.value = header
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = header_alignment
        celda.border = thin_border

    # Datos
    for row_num, empleado in enumerate(empleados, start=9):
        ws.cell(row=row_num, column=1, value=empleado.get('id', ''))
        ws.cell(row=row_num, column=2, value=empleado.get('cedula', ''))
        ws.cell(row=row_num, column=3, value=empleado.get('nombre', ''))
        ws.cell(row=row_num, column=4, value=empleado.get('apellido', ''))
        ws.cell(row=row_num, column=5, value=empleado.get('departamento', ''))
        estado = "Activo" if empleado.get('boolValidacion') == 1 else "Inactivo"
        ws.cell(row=row_num, column=6, value=estado)

    # Ajustar columnas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

    ws.column_dimensions['A'].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='reporte_empleados.xlsx')


# ============ FUNCIÓN PARA REPORTE DE BENEFICIOS ============
def generar_reporte_beneficios_excel(reporte, filtros=None):
    """
    Genera un archivo Excel con los resultados de beneficios filtrados.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Beneficios"

    # Estilos
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0a58ca', end_color='0a58ca', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    membrete_font = Font(name='Arial', size=8, bold=False, color='444444')
    membrete_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ============ LOGO ============
    ruta_logo = os.path.join('static', 'images', 'Imagen1.jpg')
    if os.path.exists(ruta_logo):
        try:
            img = Image(ruta_logo)
            img.width = 60
            img.height = 60
            ws.add_image(img, 'A1')
        except:
            pass

    # ============ MEMBRETE ============
    ws.merge_cells('B1:G1')
    ws['B1'].value = "L.P. LIDER POLLO, C.A. RIF J-30713528-0"
    ws['B1'].font = membrete_font
    ws['B1'].alignment = membrete_alignment
    
    ws.merge_cells('B2:G2')
    ws['B2'].value = "Zona Industrial El Recreo Parcela Nro. 51 Vía Flor Amarillo, Valencia Edo. Carabobo"
    ws['B2'].font = membrete_font
    ws['B2'].alignment = membrete_alignment
    
    ws.merge_cells('B3:G3')
    ws['B3'].value = "Telf: (0412) 801.7687 / 801.7887 - liderpolloca@gmail.com"
    ws['B3'].font = membrete_font
    ws['B3'].alignment = membrete_alignment
    
    ws.row_dimensions[4].height = 10

    # Título
    ws.merge_cells('A5:G5')
    ws['A5'].value = "REPORTE DE BENEFICIOS - SICEB"
    ws['A5'].font = Font(name='Arial', size=14, bold=True)
    ws['A5'].fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
    ws['A5'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    ws['A5'].alignment = Alignment(horizontal='center')
    
    # Fecha
    ws.merge_cells('A6:G6')
    ws['A6'].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A6'].font = Font(size=8)
    ws['A6'].alignment = Alignment(horizontal='center')
    
    # Encabezados
    headers = ['Empleado', 'Cédula', 'Departamento', 'Estatus', 'Productos', 'Fecha', 'Semana/Tipo']
    for col_num, header in enumerate(headers, 1):
        celda = ws.cell(row=8, column=col_num)
        celda.value = header
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = header_alignment
        celda.border = thin_border

    # Datos
    for row_num, item in enumerate(reporte, start=9):
        ws.cell(row=row_num, column=1, value=item.get('nombre_completo', ''))
        ws.cell(row=row_num, column=2, value=item.get('cedula', ''))
        ws.cell(row=row_num, column=3, value=item.get('departamento', ''))
        
        estado = "Recibió" if item.get('recibio') else "No recibió"
        ws.cell(row=row_num, column=4, value=estado)
        
        productos_list = item.get('productos_list', [])
        productos = ', '.join([f"{p[1]} {p[0]}" for p in productos_list]) if productos_list else ''
        ws.cell(row=row_num, column=5, value=productos)
        
        fecha_entrega = item.get('fecha_entrega')
        if fecha_entrega:
            if hasattr(fecha_entrega, 'strftime'):
                ws.cell(row=row_num, column=6, value=fecha_entrega.strftime('%d/%m/%Y %H:%M'))
            else:
                ws.cell(row=row_num, column=6, value=str(fecha_entrega))
        
        semana_tipo = f"S{item.get('semana')} - {item.get('tipo_nomina')}"
        ws.cell(row=row_num, column=7, value=semana_tipo)
    
    # Ajustar columnas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='reporte_beneficios.xlsx')