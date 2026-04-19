# services/empleado_excel.py
from flask import send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO
import os
from PIL import Image as PILImage
from datetime import datetime

def generar_reporte_empleados_excel(empleados):
    """
    Genera un archivo Excel personalizado con la lista de empleados
    Mismo formato que los reportes de beneficios e histórico
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Empleados"

    # ============ ESTILOS ============
    titulo_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    titulo_fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
    titulo_alignment = Alignment(horizontal='center', vertical='center')

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0a58ca', end_color='0a58ca', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    data_alignment_left = Alignment(horizontal='left', vertical='center')
    data_alignment_center = Alignment(horizontal='center', vertical='center')
    data_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Estilos para estado Activo/Inactivo (como los de Recibió/No Recibió)
    active_fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
    active_font = Font(color='155724', bold=True)
    inactive_fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
    inactive_font = Font(color='721c24', bold=True)
    
    membrete_font = Font(name='Arial', size=8, bold=False, color='444444')
    membrete_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ============ AGREGAR LOGO ============
    logo_encontrado = None
    formatos = ['png', 'PNG', 'webp', 'WEBP', 'jpg', 'JPG', 'jpeg', 'JPEG']
    
    for formato in formatos:
        ruta_logo = os.path.join('static', 'images', f'imagen1.{formato}')
        if os.path.exists(ruta_logo):
            logo_encontrado = ruta_logo
            break
    
    if not logo_encontrado:
        import glob
        archivos = glob.glob(os.path.join('static', 'images', 'imagen1.*'))
        if archivos:
            logo_encontrado = archivos[0]
    
    if logo_encontrado and os.path.exists(logo_encontrado):
        try:
            if logo_encontrado.endswith('.webp'):
                img_pil = PILImage.open(logo_encontrado)
                temp_png = BytesIO()
                img_pil.save(temp_png, format='PNG')
                temp_png.seek(0)
                img = Image(temp_png)
            else:
                img = Image(logo_encontrado)
            
            img.width = 60
            img.height = 60
            ws.add_image(img, 'A1')
        except Exception as e:
            print(f"Error al cargar el logo: {e}")
    
    # ============ MEMBRETE ============
    ws.merge_cells('B1:F1')
    celda_membrete = ws['B1']
    celda_membrete.value = "L.P. LIDER POLLO, C.A. RIF J-30713528-0"
    celda_membrete.font = membrete_font
    celda_membrete.alignment = membrete_alignment
    
    ws.merge_cells('B2:F2')
    celda_direccion = ws['B2']
    celda_direccion.value = "Zona Industrial El Recreo Parcela Nro. 51 Vía Flor Amarillo, Valencia Edo. Carabobo"
    celda_direccion.font = membrete_font
    celda_direccion.alignment = membrete_alignment
    
    ws.merge_cells('B3:F3')
    celda_contacto = ws['B3']
    celda_contacto.value = "Telf: (0412) 801.7687 / 801.7887 - liderpolloca@gmail.com"
    celda_contacto.font = membrete_font
    celda_contacto.alignment = membrete_alignment
    
    ws.row_dimensions[4].height = 10

    # ============ TÍTULO ============
    ws.merge_cells('A5:F5')
    celda_titulo = ws['A5']
    celda_titulo.value = "REPORTE DE EMPLEADOS - SICEB"
    celda_titulo.font = titulo_font
    celda_titulo.fill = titulo_fill
    celda_titulo.alignment = titulo_alignment
    ws.row_dimensions[5].height = 30

    # ============ SUBTÍTULO ============
    ws.merge_cells('A6:F6')
    celda_subtitulo = ws['A6']
    fecha_reporte = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    celda_subtitulo.value = f"Generado: {fecha_reporte}"
    celda_subtitulo.font = Font(size=9, italic=True, color='666666')
    celda_subtitulo.alignment = Alignment(horizontal='center')
    ws.row_dimensions[6].height = 20

    # ============ ENCABEZADOS ============
    headers = ['ID', 'Cédula', 'Nombre Completo', 'Departamento', 'Estado']
    for col_num, header in enumerate(headers, 1):
        celda = ws.cell(row=8, column=col_num)
        celda.value = header
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = header_alignment
        celda.border = thin_border

    # ============ DATOS ============
    for row_num, empleado in enumerate(empleados, start=9):
        # ID
        celda = ws.cell(row=row_num, column=1, value=empleado.get('id', ''))
        celda.alignment = data_alignment_center
        celda.border = data_border
        
        # Cédula
        celda = ws.cell(row=row_num, column=2, value=empleado.get('cedula', ''))
        celda.alignment = data_alignment_center
        celda.border = data_border
        
        # Nombre Completo
        nombre_completo = f"{empleado.get('nombre', '')} {empleado.get('apellido', '')}".strip()
        celda = ws.cell(row=row_num, column=3, value=nombre_completo)
        celda.alignment = data_alignment_left
        celda.border = data_border
        
        # Departamento
        celda = ws.cell(row=row_num, column=4, value=empleado.get('departamento', 'Sin asignar'))
        celda.alignment = data_alignment_left
        celda.border = data_border
        
        # Estado (Activo/Inactivo) con colores
        bool_validacion = empleado.get('boolValidacion')
        estado = "ACTIVO" if bool_validacion == 1 else "INACTIVO"
        celda = ws.cell(row=row_num, column=5, value=estado)
        celda.alignment = data_alignment_center
        celda.border = data_border
        
        if bool_validacion == 1:
            celda.fill = active_fill
            celda.font = active_font
        else:
            celda.fill = inactive_fill
            celda.font = inactive_font

    # ============ AJUSTAR ANCHO DE COLUMNAS ============
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12

    # ============ RESUMEN FINAL ============
    activos = sum(1 for e in empleados if e.get('boolValidacion') == 1)
    inactivos = sum(1 for e in empleados if e.get('boolValidacion') != 1)
    
    ultima_fila = len(empleados) + 9
    ws.merge_cells(f'A{ultima_fila + 1}:F{ultima_fila + 1}')
    celda_total = ws.cell(row=ultima_fila + 1, column=1)
    celda_total.value = f"📊 RESUMEN: Total Empleados: {len(empleados)} | Activos: {activos} | Inactivos: {inactivos}"
    celda_total.font = Font(bold=True, size=11)
    celda_total.alignment = Alignment(horizontal='center')
    celda_total.fill = PatternFill(start_color='e7f3ff', end_color='e7f3ff', fill_type='solid')

    # ============ GUARDAR ============
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'reporte_empleados_{timestamp}.xlsx'

    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)